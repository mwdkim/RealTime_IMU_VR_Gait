# MVN MXTP02 UDP → Right knee (RightUpLeg + RightLeg) → Knee ISA Plücker (Λ, Π) + pitch h → Excel
# Output columns (7): Λx Λy Λz Πx Πy Πz h

import socket, struct, time, math
import numpy as np
from openpyxl import Workbook, load_workbook

# =========================
# USER SETTINGS
# =========================
UDP_IP   = "127.0.0.1"
UDP_PORT = 9763

THIGH_SEG = "RightUpLeg"
SHANK_SEG = "RightLeg"

OUTPUT_XLSX = "plucker_right_knee.xlsx"
FLUSH_EVERY_N = 1
PRINT_EVERY_N = 10
MAX_STEPS = 300

EPS = 1e-9
OMEGA_MIN = 0.05

R_MAP = np.eye(3)

# Knee center calibration
CALIB_FRAMES = 120
calib_count = 0
knee_ready = False

o_t = None
o_s = None
# =========================
# MXTP02 PARSER (unchanged)
# =========================
HEADER_FMT   = ">6s I B B I B B B B H H"      # 24 bytes
ITEM_FMT_POS = ">i f f f f f f f"            # 32 bytes
ITEM_FMT_VEL = ">i f f f f f f f f f f"      # 44 bytes

SEG_NAMES = [
    "Pelvis","L5","L3","T12","T8","Neck","Head",
    "RightShoulder","RightUpperArm","RightForeArm","RightHand",
    "LeftShoulder","LeftUpperArm","LeftForeArm","LeftHand",
    "RightUpLeg","RightLeg","RightFoot","RightToe",
    "LeftUpLeg","LeftLeg","LeftFoot","LeftToe"
]

def seg_name_from_id(i: int):
    return SEG_NAMES[i-1] if 0 < i <= len(SEG_NAMES) else None

def parse_mxtp02(packet: bytes):
    """Return dicts: quats{name:(w,x,y,z)}, pos{name:3}, vel{name:3 (optional)}"""
    if len(packet) < 24:
        return {}, {}, {}
    try:
        idstr, sc, dg, num_items, tc, char_id, nbody, nprops, nfingers, _r, _p = struct.unpack(HEADER_FMT, packet[:24])
    except struct.error:
        return {}, {}, {}
    if not idstr.startswith(b"MXTP02"):
        return {}, {}, {}

    offset = 24
    remain = len(packet) - offset
    if num_items > 0:
        guess = remain // num_items
        item_sz = 44 if guess == 44 else 32
    else:
        item_sz = 44 if remain % 44 == 0 else 32

    quats, pos, vel = {}, {}, {}
    count = remain // item_sz
    for i in range(count):
        s = offset + i * item_sz
        try:
            if item_sz == 44:
                seg_id, px,py,pz, qw,qx,qy,qz, vx,vy,vz = struct.unpack(ITEM_FMT_VEL, packet[s:s+item_sz])
            else:
                seg_id, px,py,pz, qw,qx,qy,qz = struct.unpack(ITEM_FMT_POS, packet[s:s+item_sz])
                vx = vy = vz = None
        except struct.error:
            continue

        name = seg_name_from_id(seg_id)
        if not name:
            continue

        p = R_MAP @ np.array([px,py,pz], dtype=float)
        pos[name] = p

        if vx is not None:
            v = R_MAP @ np.array([vx,vy,vz], dtype=float)
            vel[name] = v

        quats[name] = np.array([qw,qx,qy,qz], dtype=float)

    return quats, pos, vel

# =========================
# QUATERNION + SCREW MATH (unchanged)
# =========================
def quat_normalize(q):
    n = np.linalg.norm(q)
    return q / n if n > 0 else q

def quat_conj(q):
    return np.array([q[0], -q[1], -q[2], -q[3]], dtype=float)

def quat_mul(q1, q2):
    w1,x1,y1,z1 = q1
    w2,x2,y2,z2 = q2
    return np.array([
        w1*w2 - x1*x2 - y1*y2 - z1*z2,
        w1*x2 + x1*w2 + y1*z2 - z1*y2,
        w1*y2 - x1*z2 + y1*w2 + z1*x2,
        w1*z2 + x1*y2 - y1*x2 + z1*w2
    ], dtype=float)

def quat_dot(q1, q2):
    return float(np.dot(q1, q2))

def quat_to_omega(q_prev, q_curr, dt):
    if dt <= 1e-8:
        return np.zeros(3)

    q_prev = quat_normalize(q_prev)
    q_curr = quat_normalize(q_curr)

    if quat_dot(q_prev, q_curr) < 0:
        q_curr = -q_curr

    dq = quat_mul(q_curr, quat_conj(q_prev))
    dq = quat_normalize(dq)

    w = max(-1.0, min(1.0, float(dq[0])))
    angle = 2.0 * math.acos(w)
    s = math.sqrt(max(0.0, 1.0 - w*w))
    if s < 1e-10 or angle < 1e-10:
        return np.zeros(3)

    axis = dq[1:4] / s
    return axis * (angle / dt)

def compute_screw_from_point_motion(r, v, omega, eps=1e-9):
    om2 = float(np.dot(omega, omega))
    if om2 < eps:
        return np.array([0.0,0.0,1.0]), 0.0, r.copy()

    om_norm = math.sqrt(om2)
    e = omega / (om_norm + eps)
    h = float(np.dot(v, omega)) / (om2 + eps)
    r0 = r + np.cross(e, np.cross(v, e)) / (om2 + eps)
    return e, h, r0

def plucker_from_r0_e(r0, e):
    Lam = e
    Pi = np.cross(r0, e)
    return Lam, Pi

def transport_velocity(v_A, omega, r_K, r_A):
    # v_K = v_A + omega x (r_K - r_A)
    return v_A + np.cross(omega, (r_K - r_A))

def quat_to_rotmat(q):
    q = quat_normalize(q)
    w, x, y, z = q

    return np.array([
        [1-2*(y*y+z*z), 2*(x*y-z*w), 2*(x*z+y*w)],
        [2*(x*y+z*w), 1-2*(x*x+z*z), 2*(y*z-x*w)],
        [2*(x*z-y*w), 2*(y*z+x*w), 1-2*(x*x+y*y)]
    ])
# =========================
# EXCEL WRITER
# =========================
HEADERS = ["Lx","Ly","Lz","Pix","Piy","Piz","h"]

def init_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "plucker"
    ws.append(HEADERS)
    wb.save(path)

def append_rows(path, rows):
    wb = load_workbook(path)
    ws = wb["plucker"]
    for r in rows:
        ws.append(r)
    wb.save(path)

# =========================
# MAIN UDP LOOP
# =========================
sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
sock.bind((UDP_IP, UDP_PORT))
sock.settimeout(1.0)

init_workbook(OUTPUT_XLSX)
print(f"Listening UDP on {UDP_IP}:{UDP_PORT}")
print(f"Tracking knee using: {THIGH_SEG} (thigh) + {SHANK_SEG} (shank)")
print(f"Writing Excel: {OUTPUT_XLSX}")
print(f"Will stop after {MAX_STEPS} steps.\n")

# Previous states for both segments
prev = {
    THIGH_SEG: {"q": None, "r": None, "t": None},
    SHANK_SEG: {"q": None, "r": None, "t": None},
}

buffer = []
total = 0

try:
    while total < MAX_STEPS:
        try:
            packet, _addr = sock.recvfrom(4096)
        except socket.timeout:
            continue

        t = time.time()
        quats, pos, vel = parse_mxtp02(packet)

        if THIGH_SEG not in quats or THIGH_SEG not in pos:
            continue
        if SHANK_SEG not in quats or SHANK_SEG not in pos:
            continue

        # Current samples
        q_t = quat_normalize(quats[THIGH_SEG])
        r_t = pos[THIGH_SEG]
        v_t = vel.get(THIGH_SEG, None)

        q_s = quat_normalize(quats[SHANK_SEG])
        r_s = pos[SHANK_SEG]
        v_s = vel.get(SHANK_SEG, None)

        # Initialize
        if prev[THIGH_SEG]["t"] is None or prev[SHANK_SEG]["t"] is None:
            prev[THIGH_SEG].update({"q": q_t, "r": r_t, "t": t})
            prev[SHANK_SEG].update({"q": q_s, "r": r_s, "t": t})
            continue

        dt = max(1e-4, min(0.05, t - prev[THIGH_SEG]["t"]))

        # Angular velocities from quaternion difference
        omega_t = quat_to_omega(prev[THIGH_SEG]["q"], q_t, dt)
        omega_s = quat_to_omega(prev[SHANK_SEG]["q"], q_s, dt)

        # Linear velocities at each segment origin
        if v_t is None:
            v_tA = (r_t - prev[THIGH_SEG]["r"]) / dt
        else:
            v_tA = v_t

        if v_s is None:
            v_sA = (r_s - prev[SHANK_SEG]["r"]) / dt
        else:
            v_sA = v_s

                # ----------------------------------------
        # Knee center estimation (fixed-offset)
        # ----------------------------------------

     

        if not knee_ready:

            R_t = quat_to_rotmat(q_t)
            R_s = quat_to_rotmat(q_s)

            r_K0 = 0.5*(r_t + r_s)

            o_t = R_t.T @ (r_K0 - r_t)
            o_s = R_s.T @ (r_K0 - r_s)

            calib_count += 1

            if calib_count >= CALIB_FRAMES:
                knee_ready = True
                print("Knee center calibrated.")

            r_K = r_K0

        else:

            R_t = quat_to_rotmat(q_t)
            R_s = quat_to_rotmat(q_s)

            r_K_t = r_t + R_t @ o_t
            r_K_s = r_s + R_s @ o_s

            r_K = 0.5*(r_K_t + r_K_s)

        # Transport both velocities to knee center
        v_tK = transport_velocity(v_tA, omega_t, r_K, r_t)
        v_sK = transport_velocity(v_sA, omega_s, r_K, r_s)

        # Relative knee twist
        omega_k = omega_s - omega_t
        v_k = v_sK - v_tK

        # Knee ISA from relative twist (reference point = r_K)
        e, h, r0 = compute_screw_from_point_motion(r_K, v_k, omega_k, eps=EPS)

        # Guard near-zero omega
        if float(np.dot(omega_k, omega_k)) < (OMEGA_MIN**2):
            h = 0.0

        Lam, Pi = plucker_from_r0_e(r0, e)

        row = [float(Lam[0]), float(Lam[1]), float(Lam[2]),
               float(Pi[0]),  float(Pi[1]),  float(Pi[2]),
               float(h)]
        buffer.append(row)
        total += 1

        if total % PRINT_EVERY_N == 0:
            print(f"[{total:3d}/{MAX_STEPS}] knee pitch h = {h:+.6f} | |ω_knee| = {float(np.linalg.norm(omega_k)):.4f} rad/s")

        if len(buffer) >= FLUSH_EVERY_N:
            append_rows(OUTPUT_XLSX, buffer)
            buffer = []

        # Update prev
        prev[THIGH_SEG].update({"q": q_t, "r": r_t, "t": t})
        prev[SHANK_SEG].update({"q": q_s, "r": r_s, "t": t})

except KeyboardInterrupt:
    print("\nStopped by user.")

finally:
    sock.close()
    if buffer:
        append_rows(OUTPUT_XLSX, buffer)
    print(f"Done. Saved {total} rows → {OUTPUT_XLSX}")